import pandas as pd
import os
import json
import logging
import asyncio
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from pydantic import BaseModel, Field
from langchain_core.tools import tool
from config import settings

logger = logging.getLogger(__name__)

# --- SCHEMAS EXPLICITOS ---

class ReadExcelSchema(BaseModel):
    file_path: Optional[str] = Field(default=None, description="O nome do arquivo Excel (ex: 'dados.xlsx') a ser lido.")
    sheet_name: Union[str, int, None] = Field(default=0, description="Opcional. Nome da aba ou índice (0 para a primeira).")

class WriteExcelSchema(BaseModel):
    data: Optional[Any] = Field(default=None, description="OBRIGATÓRIO. Os dados para criar a planilha.")
    file_path: Optional[str] = Field(default=None, description="OBRIGATÓRIO. O nome do arquivo Excel (ex: 'relatorio.xlsx').")
    sheet_name: Optional[str] = Field(default="Sheet1", description="Opcional. Nome da aba.")

def _clean_data(raw_data: Any) -> List[Dict[str, Any]]:
    """Tenta limpar e consertar o payload de dados enviado pelo modelo Gemini/OpenAI."""
    logger.info(f"📊 [ExcelTool] Recebido para limpeza: tipo={type(raw_data)}")
    
    if raw_data is None:
        logger.error("[ExcelTool] Erro: Dados recebidos como None.")
        raise ValueError("O campo 'data' foi recebido como None/Empty. Você deve prover dados para criar uma planilha.")
        
    if isinstance(raw_data, str):
        try:
            raw_data = json.loads(raw_data)
        except Exception:
            raise ValueError("O campo 'data' foi enviado como uma string mas não é um JSON válido.")
            
    if not isinstance(raw_data, list):
        if isinstance(raw_data, dict):
            raw_data = [raw_data]
        else:
            raise ValueError(f"O campo 'data' deve ser uma lista (array) de dicionários (objetos). Recebeu: {type(raw_data)}")
            
    cleaned = []
    for item in raw_data:
        if isinstance(item, dict):
            cleaned.append(item)
        elif isinstance(item, list):
            # Converte lista em colunas numeradas
            cleaned.append({f"Coluna_{i+1}": val for i, val in enumerate(item)})
        else:
            cleaned.append({"Valor": str(item)})
            
    if not cleaned:
        raise ValueError("A lista de dados estava vazia após a limpeza.")
        
    return cleaned

def _apply_premium_style(file_path: str):
    """Aplica o design corporativo premium (Manus AI) à planilha gerada."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path)
        
        # Paleta Corporativa Premium Executiva
        header_fill = PatternFill(start_color="1C4E9E", end_color="1C4E9E", fill_type="solid") # Azul Corporativo
        zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Cinza super claro
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
        cell_font = Font(name="Calibri", size=11, color="333333")
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'), 
            right=Side(style='thin', color='E5E7EB'), 
            top=Side(style='thin', color='E5E7EB'), 
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Congelar painel superior (Header) e adicionar Filtro
            ws.freeze_panes = "A2"
            if ws.max_column > 0:
                end_col_letter = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{end_col_letter}{ws.max_row}"
            
            # Formatar Cabeçalhos
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Formatar Células, Stripe e Ajustar Largura
            for col in range(1, ws.max_column + 1):
                max_len = len(str(ws.cell(row=1, column=col).value or ""))
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    c.font = cell_font
                    c.border = thin_border
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                    
                    # Zebra Striping (linhas pares com fundo leve)
                    if row % 2 == 0:
                        c.fill = zebra_fill

                    if c.value:
                        max_len = max(max_len, len(str(c.value)))
                
                # Ajuste inteligente de largura de coluna (limitado a 60 caracteres)
                adjusted_width = min(max_len + 5, 80)
                ws.column_dimensions[get_column_letter(col)].width = adjusted_width

        wb.save(file_path)
        logger.info(f"✨ [ExcelGen] Formatação Executiva aplicada com sucesso em: {file_path}")
    except Exception as e:
        logger.warning(f"⚠️ [ExcelGen] Falhou ao aplicar formatação premium em {file_path}: {e}")

@tool(args_schema=ReadExcelSchema)
async def read_excel(file_path: str, sheet_name: Union[str, int] = 0) -> List[Dict[str, Any]]:
    """
    Lê uma planilha Excel e retorna os dados como uma lista de dicionários.
    Ideal para recuperar dados previamente gerados ou analisar planilhas existentes.
    """
    if file_path is None:
        logger.error("[ExcelTool] Erro crítico: file_path é None!")
        return [{"error": "file_path não pode ser nulo"}]
        
    try:
        if not os.path.isabs(file_path):
            file_path = os.path.join(settings.DATA_OUTPUTS_PATH, file_path)
            
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
            
        df = await asyncio.to_thread(pd.read_excel, file_path, sheet_name=sheet_name)
        data = df.to_dict(orient="records")
        logger.info(f"Sucesso ao ler Excel: {file_path} ({len(data)} linhas)")
        return data
    except Exception as e:
        logger.error(f"Erro ao ler Excel {file_path}: {e}")
        return [{"error": str(e)}]

@tool(args_schema=WriteExcelSchema)
async def create_excel(data: list, file_path: str, sheet_name: str = "Sheet1") -> str:
    """
    Cria UMA NOVA planilha Excel (.xlsx) inteira a partir de uma lista de dados JSON.
    Esta ferramenta deve ser usada para consolidar relatórios, tabelas de tendências ou novos documentos.
    O campo 'data' deve ser OBRIGATORIAMENTE uma lista (array).
    """
    if data is None or file_path is None:
        logger.error(f"[ExcelGen] ERRO CRÍTICO: Parâmetros nulos! data={data}, file_path={file_path}")
        if data is None: data = []
        if file_path is None: file_path = f"relatorio_{int(datetime.now().timestamp())}.xlsx"

    try:
        logger.info(f"🚀 [ExcelGen] Iniciando criação de {file_path}...")
        clean_data = _clean_data(data)
        
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"
            
        full_path = os.path.join(settings.DATA_OUTPUTS_PATH, file_path)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        # --- TENTATIVA 1: PANDAS ---
        try:
            logger.info("[ExcelGen] Tentando via PANDAS...")
            df = pd.DataFrame(clean_data)
            await asyncio.to_thread(df.to_excel, full_path, index=False, sheet_name=sheet_name)
            await asyncio.to_thread(_apply_premium_style, full_path)
            logger.info(f"✅ [ExcelGen] Sucesso via PANDAS: {full_path}")
            size = os.path.getsize(full_path)
            return f"Planilha '{file_path}' gerada com sucesso ({size} bytes). <SEND_FILE:{os.path.basename(full_path)}>"
        except Exception as pe:
            logger.warning(f"⚠️ [ExcelGen] Pandas falhou: {pe}. Tentando FALLBACK OpenPyXL...")
            
            # --- TENTATIVA 2: OPENPYXL NATIVO ---
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if clean_data:
                headers = list(clean_data[0].keys())
                ws.append(headers)
                for row in clean_data:
                    ws.append([row.get(h, "") for h in headers])
            
            await asyncio.to_thread(wb.save, full_path)
            await asyncio.to_thread(_apply_premium_style, full_path)
            logger.info(f"✅ [ExcelGen] Sucesso via FALLBACK OpenPyXL: {full_path}")
            return f"Planilha '{file_path}' criada via fallback seguro com {len(clean_data)} linhas. <SEND_FILE:{os.path.basename(full_path)}>"
            
    except Exception as e:
        logger.error(f"❌ [ExcelGen] Erro crítico ao criar Excel {file_path}: {e}")
        return f"Erro ao criar planilha: {str(e)}"

@tool(args_schema=WriteExcelSchema)
async def append_to_excel(data: list, file_path: str, sheet_name: str = "Sheet1") -> str:
    """
    Adiciona NOVAS LINHAS a uma planilha Excel já existente de forma incremental.
    """
    if data is None:
        logger.error("[ExcelTool] Erro: append_to_excel recebeu data=None. Usando lista vazia.")
        data = []
        
    try:
        clean_data = _clean_data(data)
        
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"
            
        if not os.path.isabs(file_path):
            full_path = os.path.join(settings.DATA_OUTPUTS_PATH, file_path)
        else:
            full_path = file_path
            
        if not os.path.exists(full_path):
            return await create_excel.ainvoke({"data": clean_data, "file_path": file_path, "sheet_name": sheet_name})
            
        existing_df = await asyncio.to_thread(pd.read_excel, full_path, sheet_name=sheet_name)
        new_df = pd.DataFrame(clean_data)
        
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        await asyncio.to_thread(combined_df.to_excel, full_path, index=False, sheet_name=sheet_name)
        
        logger.info(f"Dados adicionados ao Excel: {full_path}")
        return f"Mais {len(clean_data)} linhas adicionadas à planilha '{os.path.basename(full_path)}'. <SEND_FILE:{os.path.basename(full_path)}>"
    except Exception as e:
        logger.error(f"Erro ao atualizar Excel {file_path}: {e}")
        return f"Erro ao atualizar planilha: {str(e)}"
